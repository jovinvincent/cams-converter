"""
CAMS CAS → Portfolio Excel Converter
One-click web app: upload PDF, get Excel automatically.
Free hosting on Streamlit Community Cloud.
Uses the SAME proven parser as qsif_pipeline.py.
"""

import io
import re
import pdfplumber
import streamlit as st
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# CONFIG
# ============================================================================

QUANT_SIF_ISINS_ORDERED = [
    "INF966L30019",  # Equity Long-Short Fund
    "INF966L30159",  # Equity Ex-Top 100 Long-Short Fund
    "INF966L30274",  # Sector Rotation Long-Short Fund
    "INF966L30241",  # Active Asset Allocator Long-Short Fund
    "INF966L30076",  # Hybrid Long-Short Fund
]
QUANT_SIF_ISINS = set(QUANT_SIF_ISINS_ORDERED)

QUANT_SIF_NAMES = {
    "INF966L30019": "Equity Long-Short Fund",
    "INF966L30159": "Equity Ex-Top 100 Long-Short",
    "INF966L30274": "Sector Rotation LongShort Fu",
    "INF966L30241": "Active Asset Allocator Long-",
    "INF966L30076": "Hybrid Long-Short Fund",
}
# Full fund names (used in title and column C) — not truncated
QUANT_SIF_FULL_NAMES = {
    "INF966L30019": "Equity Long-Short Fund",
    "INF966L30159": "Equity Ex-Top 100 Long-Short Fund Direct Plan",
    "INF966L30274": "Sector Rotation LongShort Fund",
    "INF966L30241": "Active Asset Allocator Long-Short Fund Direct Plan",
    "INF966L30076": "Hybrid Long-Short Fund",
}
LIQUID_FUND_ISIN = "INF966L01820"

LATEST_NAV = {
    "INF966L30019": 10.9871,
    "INF966L30159": 10.7319,
    "INF966L30274": 10.3595,
    "INF966L30241": 10.7997,
    "INF966L30076": 10.7412,
}

STAMP_DUTY_RATE = 0.00005
STT_RATE = 0.00001
# CAMS statement date (used for holding days calculation)
STATEMENT_DATE = datetime(2026, 7, 8)
XIRR_VALUATION_DATE = datetime(2026, 7, 10)  # last trading day (for XIRR only)

COMMON_PASSWORDS = ["jovi31490", "", "IPRU9999", "CAMSCAS", "CAS123", "12345678", "password"]


# ============================================================================
# PDF PARSING (mirrors qsif_pipeline.py exactly)
# ============================================================================

def extract_pdf_text(pdf_bytes: bytes, password: str = "") -> str:
    passwords_to_try = [password] if password else []
    passwords_to_try.extend(p for p in COMMON_PASSWORDS if p != password)

    last_err = None
    for pwd in passwords_to_try:
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes), password=pwd or None) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)
        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(
        f"Could not open PDF with any password. Tried: {passwords_to_try}. "
        f"Last error: {last_err}"
    )


def parse_num(s: str) -> float:
    s = s.replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_cas_pdf(full_text: str) -> dict:
    """
    Parse CAMS CAS text into {isin: fund_dict}. Charge lines (***) are folded
    into the preceding non-charge transaction's STT / Stamp Duty / TDS columns.
    Uses the SAME logic as qsif_pipeline.py parse_funds().
    """
    fund_blocks = re.split(r"(?=Folio No:)", full_text)
    funds = {}

    for block in fund_blocks[1:]:
        isin_m = re.search(r"ISIN:\s*(\S+)", block)
        if not isin_m:
            continue
        isin = isin_m.group(1)
        if isin not in QUANT_SIF_ISINS:
            continue

        # Closing CAMS line
        closing = re.search(
            r"Closing Unit Balance:\s*([\d,]+\.\d+).*?NAV on\s*\d{2}-[A-Za-z]{3}-\d{4}:\s*INR\s*([\d,]+\.\d+).*?"
            r"Total Cost Value:\s*([\d,]+\.\d+)",
            block, re.DOTALL,
        )
        if closing:
            remaining_units = float(closing.group(1).replace(",", ""))
            cams_closing_nav = float(closing.group(2).replace(",", ""))
            cams_cost = float(closing.group(3).replace(",", ""))
        else:
            remaining_units = 0.0
            cams_closing_nav = None
            cams_cost = 0.0

        # Walk lines; fold charge lines into the last non-charge transaction
        transactions = []
        last_idx = -1
        for raw in block.split("\n"):
            line = raw.strip()
            if re.match(r"\d{2}-[A-Z][a-z]{2}-\d{4}\s+To\s+\d{2}-[A-Z][a-z]{2}-\d{4}", line):
                continue
            if "Address Updated from CVL Data" in line:
                continue
            if re.match(r"(Nominee|Opening Unit Balance|Closing Unit Balance|Total Cost Value|Market Value|NAV on|Entry Load|Exit Load|Graded Exit|Folio No|PAN:|KYC:|Registrar|Statement Date|This Consolidated|This statement|This CAS|Email Id|Mobile|investor friendly|registered|common to|missing from|consolidate all|check with|reverse|brought to|of your family|list the|Period:|^\W*$|JOVIN VINCENT|Address:|KERALA|KOTTAYAM|INDIA|PLASSANAL|MEENACHIL|THALAPPALAM|S/O VINCENT JOSEPH|OLAYATHIL|PAN Number|S\.O\.)", line):
                continue

            date_match = re.match(r"(\d{2}-[A-Z][a-z]{2}-\d{4})\s+(.+)", line)
            if not date_match:
                continue
            date_str = date_match.group(1)
            rest = date_match.group(2)

            # Charge line
            charge_m = re.search(r"\*+\s*([^*]+?)\s*\*+\s*\(?(-?[\d,]+\.?\d*)\)?\s*$", rest)
            if charge_m and "***" in rest:
                charge_type = charge_m.group(1).strip()
                try:
                    amount = float(charge_m.group(2).replace(",", "").replace("(", "-").replace(")", ""))
                except ValueError:
                    amount = 0.0
                if last_idx >= 0:
                    ct = charge_type.lower()
                    if "stt" in ct:
                        transactions[last_idx]["stt"] = amount
                    elif "tds" in ct:
                        transactions[last_idx]["tds"] = amount
                    elif "stamp" in ct:
                        transactions[last_idx]["stamp_duty"] = amount
                continue

            # Regular transaction
            clean = re.sub(r"\(NAV Dt\s*:\s*\d{2}/\d{2}/\d{4}\)", "", rest)
            clean = re.sub(r"\s+less\s+(TDS|STT|TDS,\s*STT|TDS,STT)", "", clean)
            nums = re.findall(r"\(?-?[\d,]+\.\d+\)?", clean)
            if len(nums) >= 4:
                amount = parse_num(nums[-4])
                units = parse_num(nums[-3])
                nav = parse_num(nums[-2])
                balance = parse_num(nums[-1])
                first_num_pos = clean.find(nums[0])
                ttype = clean[:first_num_pos].strip()
                ttype = re.split(r"\s*\*\*\*", ttype)[0].strip()
                ttype = re.sub(r"\s+(Trxn\.Ref\.No\..*|less.*|Inter Bank.*)$", "", ttype).strip()
                is_p = ("Purchase" in ttype or "Shift In" in ttype or "New Purchase" in ttype)
                tx = {
                    "date": date_str, "type": ttype,
                    "amount": amount, "units": units, "nav": nav, "balance": balance,
                    "stt": 0.0,
                    "stamp_duty": round(amount * STAMP_DUTY_RATE, 2) if is_p else 0.0,
                    "tds": 0.0,
                }
                transactions.append(tx)
                last_idx = len(transactions) - 1

        funds[isin] = {
            "name": QUANT_SIF_NAMES[isin],
            "isin": isin,
            "cams_cost_value": cams_cost,
            "cams_remaining_units": remaining_units,
            "cams_closing_nav": cams_closing_nav,
            "transactions": transactions,
        }
    return funds


# ============================================================================
# EXCEL WRITER (locked-in format)
# ============================================================================

def build_excel(funds: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    summary_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    summary_fill = PatternFill("solid", fgColor="2E75B6")
    purchase_font = Font(name="Calibri", size=10, color="006100")
    redeem_font = Font(name="Calibri", size=10, color="9C0006")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#", "Date", "Fund", "ISIN", "Type", "Amount (₹)", "STT 0.001% (₹)",
        "Stamp Duty (₹)", "TDS (₹)", "Total Invested (₹)", "Units", "NAV (₹)",
        "Value (₹)", "Balance Units", "Total Investor Cost (₹)", "Cost/Unit (₹)",
        "Holding Days", "Period", "Exit Load (₹)", "Notes",
    ]

    for isin in QUANT_SIF_ISINS_ORDERED:
        sheet_name = QUANT_SIF_NAMES[isin]
        full_name = QUANT_SIF_FULL_NAMES[isin]
        full_title = f"{full_name}    (ISIN: {isin})"
        # Column C: "Quant SIF - " + full name, adding " - Direct Plan" suffix only if not already present
        if full_name.endswith("Direct Plan"):
            fund_c_name = f"Quant SIF - {full_name}"
        else:
            fund_c_name = f"Quant SIF - {full_name} - Direct Plan"
        ws = wb.create_sheet(sheet_name)
        fund = funds.get(isin, {"transactions": [], "cams_cost_value": 0, "cams_remaining_units": 0})
        txns = fund["transactions"]

        # Row 1: Title (full fund name + ISIN)
        ws.merge_cells("A1:T1")
        c = ws.cell(row=1, column=1, value=full_title)
        c.font = title_font
        c.fill = title_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 28

        # Row 3: Headers
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[3].height = 32

        # Rows 4+: Transactions
        lots = []
        for i, t in enumerate(txns):
            r = 4 + i
            units = t["units"]
            txn_date = datetime.strptime(t["date"], "%d-%b-%b" if False else "%d-%b-%Y")
            if units > 0:
                # BUY: add to lots, holding days = days from txn to statement date
                stamp = t.get("stamp_duty", 0)
                cpu = (t["amount"] + stamp) / units
                lots.append({"units": units, "cpu": cpu, "date": t["date"]})
                holding_days = (STATEMENT_DATE - txn_date).days
            elif units < 0:
                # SELL: FIFO consume oldest lots first, holding days = weighted avg of consumed lots
                to_sell = abs(units)
                new_lots = []
                consumed_weighted_days = 0.0
                consumed_total = 0.0
                for lot in lots:
                    if to_sell <= 0:
                        new_lots.append(lot)
                        continue
                    take = min(lot["units"], to_sell)
                    lot_date = datetime.strptime(lot["date"], "%d-%b-%Y")
                    lot_holding = (txn_date - lot_date).days
                    consumed_weighted_days += take * lot_holding
                    consumed_total += take
                    if lot["units"] - take > 1e-9:
                        new_lots.append({"units": lot["units"] - take, "cpu": lot["cpu"], "date": lot["date"]})
                    to_sell -= take
                lots = new_lots
                # Holding days for the sell = weighted avg of consumed lots' holding periods
                # Use round() (not int()) to avoid floating-point truncation issues
                holding_days = round(consumed_weighted_days / consumed_total) if consumed_total > 0 else 0
            else:
                holding_days = 0

            balance = sum(l["units"] for l in lots)
            total_cost = sum(l["units"] * l["cpu"] for l in lots)

            ttype = t["type"]
            is_buy = ("Purchase" in ttype or "Shift In" in ttype)
            row_vals = [
                i + 1, t["date"], fund_c_name, isin, ttype,  # A-E
                t["amount"], t.get("stt", 0), t.get("stamp_duty", 0), t.get("tds", 0),  # F-I
                None,  # J (Total Invested) - empty in target format
                t["units"], t["nav"], None,  # K-M
                round(balance, 3), round(total_cost, 2),  # N-O
                None,  # P (Cost/Unit) - empty
                holding_days,  # Q (Holding Days) - days to STATEMENT for buy, FIFO lot age for sell
                None,  # R (Period) - empty
                None,  # S (Exit Load) - empty
                "Purchase" if is_buy else "Redemption net of TDS/STT",  # T (Notes)
            ]
            for col, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=col, value=v)
                c.font = purchase_font if is_buy else redeem_font
                c.alignment = Alignment(horizontal="center" if col in (1, 4, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19) else "left", vertical="center")
                c.border = border
            ws.row_dimensions[r].height = 18

        # SUMMARY block (last 4 rows)
        if txns:
            nav = LATEST_NAV.get(isin, 0)
            final_balance = sum(l["units"] for l in lots)
            final_cost = sum(l["units"] * l["cpu"] for l in lots)

            total_lot_units = sum(l["units"] for l in lots)
            if total_lot_units > 0:
                wd = sum(
                    l["units"] * (XIRR_VALUATION_DATE - datetime.strptime(l["date"], "%d-%b-%Y")).days
                    for l in lots
                ) / total_lot_units
            else:
                wd = 0
            xirr = ((nav * final_balance / final_cost) ** (365 / wd) - 1) * 100 if (wd > 0 and final_cost > 0) else 0

            n_txn = len(txns)
            sr_summary = 4 + n_txn + 1
            sr_cost = sr_summary + 1
            sr_bal = sr_summary + 2
            sr_xirr = sr_summary + 3

            for c in range(1, 21):
                ws.cell(row=sr_summary, column=c).fill = summary_fill
                ws.cell(row=sr_cost, column=c).fill = summary_fill
                ws.cell(row=sr_bal, column=c).fill = summary_fill
                ws.cell(row=sr_xirr, column=c).fill = summary_fill

            ws.cell(row=sr_summary, column=1, value="  SUMMARY").font = summary_font
            ws.cell(row=sr_summary, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

            ws.cell(row=sr_cost, column=1, value="  Total Investor Cost (₹)").font = summary_font
            ws.cell(row=sr_cost, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cc = ws.cell(row=sr_cost, column=5, value=round(final_cost, 2))
            cc.font = summary_font
            cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cc.number_format = '#,##0.00'

            ws.cell(row=sr_bal, column=1, value="  Balance Units").font = summary_font
            ws.cell(row=sr_bal, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cc = ws.cell(row=sr_bal, column=5, value=round(final_balance, 3))
            cc.font = summary_font
            cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cc.number_format = '#,##0.000'

            ws.cell(row=sr_xirr, column=1, value="  XIRR (Annualized)").font = summary_font
            ws.cell(row=sr_xirr, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cc = ws.cell(row=sr_xirr, column=5, value=None)
            cc.font = summary_font
            cc.alignment = Alignment(horizontal="right", vertical="center", indent=1)

            for r in (sr_summary, sr_cost, sr_bal, sr_xirr):
                ws.row_dimensions[r].height = 22

        widths = [5, 13, 35, 14, 22, 13, 11, 12, 9, 14, 11, 10, 11, 13, 16, 11, 11, 9, 11, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# STREAMLIT UI
# ============================================================================

st.set_page_config(
    page_title="CAMS → Excel",
    page_icon="📊",
    layout="centered",
)

st.title("📊 CAMS CAS → Portfolio Excel")
st.write("Upload your CAMS PDF. Excel downloads automatically.")

with st.expander("⚙️ PDF password (pre-filled with default — edit if needed)"):
    pdf_password = st.text_input(
        "PDF password",
        type="password",
        value="jovi31490",
        help="Default is 'jovi31490'. Edit if your PDF uses a different password.",
    )
    st.caption("💡 Default password 'jovi31490' is pre-filled. Edit only if needed.")

uploaded = st.file_uploader(
    "Choose CAMS PDF",
    type=["pdf"],
    label_visibility="visible",
)

if uploaded is not None:
    with st.spinner("Processing PDF..."):
        try:
            pdf_bytes = uploaded.read()
            text = extract_pdf_text(pdf_bytes, pdf_password)
            funds = parse_cas_pdf(text)
            xlsx_bytes = build_excel(funds)
        except Exception as e:
            st.error(f"❌ Failed: {e}")
            st.info("👉 Open the 'PDF password' section above and try entering your password (or PAN number).")
            st.stop()

    total_txn = sum(len(f["transactions"]) for f in funds.values())
    if total_txn == 0:
        st.warning("No transactions found. Check password or PDF format.")
        st.stop()

    st.success(f"✅ {total_txn} transactions parsed across 5 funds")
    st.download_button(
        label="📥 Download Excel",
        data=xlsx_bytes,
        file_name="Portfolio_Transactions.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

